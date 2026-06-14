import io
import logging
from typing import Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.services.storage_service import storage_service

logger = logging.getLogger("analytics.xlsx_exporter")


class XLSXExporter:
    async def export(self, report_data: Dict[str, Any], file_key: str) -> str:
        """Generate an Excel workbook using openpyxl and save to storage_service."""
        logger.info(f"Generating XLSX export: {file_key}")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report Metrics"

        # Styles definition
        title_font = Font(name="Segoe UI", size=16, bold=True, color="1B5E20")  # Forest Green
        section_font = Font(name="Segoe UI", size=12, bold=True)
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Segoe UI", size=11)

        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")  # Green
        summary_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # Light green tint

        thin_border = Border(
            left=Side(style="thin", color="DDDDDD"),
            right=Side(style="thin", color="DDDDDD"),
            top=Side(style="thin", color="DDDDDD"),
            bottom=Side(style="thin", color="DDDDDD"),
        )

        # Title
        ws["A1"] = f"{report_data.get('report_type', 'N/A').replace('_', ' ').upper()} REPORT"
        ws["A1"].font = title_font
        ws["A2"] = f"Generated at: {report_data.get('generated_at', '')}"
        ws["A2"].font = Font(name="Segoe UI", size=9, italic=True)

        row_idx = 4
        # Summary Header
        ws.cell(row=row_idx, column=1, value="Summary Indicators").font = section_font
        row_idx += 1

        summary = report_data.get("summary", {})
        for key, val in summary.items():
            ws.cell(row=row_idx, column=1, value=key.replace("_", " ").title()).font = Font(
                name="Segoe UI", size=11, bold=True
            )
            ws.cell(row=row_idx, column=2, value=val).font = data_font
            ws.cell(row=row_idx, column=1).fill = summary_fill
            ws.cell(row=row_idx, column=2).fill = summary_fill
            ws.cell(row=row_idx, column=1).border = thin_border
            ws.cell(row=row_idx, column=2).border = thin_border
            row_idx += 1

        row_idx += 2
        # Data Header
        ws.cell(row=row_idx, column=1, value="Detailed Records").font = section_font
        row_idx += 1

        data_list = report_data.get("data", [])
        if data_list:
            headers = list(data_list[0].keys())
            # Write headers
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=h.replace("_", " ").title())
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border
            row_idx += 1

            # Write data rows
            for row_data in data_list:
                for col_idx, h in enumerate(headers, start=1):
                    val = row_data.get(h)
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = data_font
                    cell.border = thin_border
                    # Align numbers
                    if isinstance(val, (int, float)):
                        cell.alignment = Alignment(horizontal="right")
                row_idx += 1
        else:
            ws.cell(row=row_idx, column=1, value="No records matches report filters.").font = Font(
                name="Segoe UI", size=11, italic=True
            )

        # Autofit columns
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # Save workbook to memory stream
        output = io.BytesIO()
        wb.save(output)
        saved_path = await storage_service.save_file(output.getvalue(), file_key)
        return saved_path


xlsx_exporter = XLSXExporter()
